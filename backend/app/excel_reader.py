"""
Lectura y normalización de los 3 dashboards institucionales (BGU).

Cada uno de los 3 Excel tiene una hoja de "registro" con estructura distinta.
Este módulo lee cada hoja con openpyxl (data_only=True), la normaliza a un
esquema común de KPI y expone estructuras listas para la API.

IMPORTANTE: los tres archivos se leen EN VIVO cada vez que se llama a
`load_all_data()`. No hay datos hardcodeados: si el Excel cambia, el
dashboard cambia. Se cachea en memoria y se puede forzar recarga con
`load_all_data(force_reload=True)`.
"""

from __future__ import annotations

import os
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

PATH_ESTRATEGICO = os.path.join(DATA_DIR, "plan_estrategico.xlsx")
PATH_EFECTIVIDAD = os.path.join(DATA_DIR, "plan_efectividad.xlsx")
PATH_EVALUACION = os.path.join(DATA_DIR, "plan_evaluacion_resultados.xlsx")

SHEET_ESTRATEGICO = "Registro Nuevo - Reporte 2025 -"
SHEET_EFECTIVIDAD = "Registro Nuevo - Reporte 2025 -"
SHEET_EVALUACION = "Registro - Reporte 2025 - 2026"

# ---------------------------------------------------------------------------
# Estados válidos (canónicos) tal como los pide el usuario
# ---------------------------------------------------------------------------
ESTADO_CUMPLIDO = "Cumplido"
ESTADO_PARCIAL = "Parcialmente cumplido"
ESTADO_NO_CUMPLIDO = "No cumplido"
ESTADO_SIN_DATOS = "Sin datos"
ESTADO_NO_APLICABLE = "No aplicable / Cohorte no madura"

ESTADOS_ORDEN = [
    ESTADO_CUMPLIDO,
    ESTADO_PARCIAL,
    ESTADO_NO_CUMPLIDO,
    ESTADO_SIN_DATOS,
    ESTADO_NO_APLICABLE,
]

# Colores institucionales sugeridos por estado (se usan también en el frontend
# como fallback, pero el frontend tiene su propia paleta — esto es para
# eventuales exports/PDF generados en backend)
ESTADO_COLOR = {
    ESTADO_CUMPLIDO: "#1E8F5F",
    ESTADO_PARCIAL: "#F2A93B",
    ESTADO_NO_CUMPLIDO: "#D64545",
    ESTADO_SIN_DATOS: "#9AA5B1",
    ESTADO_NO_APLICABLE: "#5B6B8C",
}


def _clean(value) -> str:
    """Normaliza texto: quita saltos de línea, espacios múltiples y trim."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _strip_accents_lower(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return text.lower().strip()


def normalize_estado(raw) -> str:
    """Mapea cualquier variante textual de Estado a uno de los 5 estados canónicos."""
    text = _strip_accents_lower(_clean(raw))
    if not text:
        return ESTADO_SIN_DATOS
    if "no aplicable" in text or "cohorte no madura" in text:
        return ESTADO_NO_APLICABLE
    if "parcial" in text:
        return ESTADO_PARCIAL
    if text.startswith("no cumplido") or text == "no cumplido":
        return ESTADO_NO_CUMPLIDO
    if "sin dato" in text:
        return ESTADO_SIN_DATOS
    if "cumplido" in text:
        return ESTADO_CUMPLIDO
    return ESTADO_SIN_DATOS


_SMALL_WORDS = {"de", "del", "y", "la", "el", "en", "a", "las", "los", "e", "al"}


def _title_case_es(text: str) -> str:
    """Title-case ligero que respeta preposiciones/artículos en español, para
    unificar la escritura de Dimensiones entre los 3 archivos (algunos vienen
    en 'Sentence case' y otros en 'Title Case')."""
    text = text.strip()
    if not text:
        return text
    words = text.split(" ")
    out = []
    for i, w in enumerate(words):
        if i > 0 and w.lower() in _SMALL_WORDS:
            out.append(w.lower())
        else:
            out.append(w[:1].upper() + w[1:].lower() if w else w)
    return " ".join(out)


def _fmt_meta_or_result(value) -> str:
    """Convierte metas/resultados numéricos (a veces fracción, a veces %) a texto legible."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Heurística: si está entre 0 y 1 probablemente es proporción -> %
        if 0 <= value <= 1:
            return f"{value * 100:.1f}%"
        return f"{value:g}"
    return _clean(value)


@dataclass
class KPI:
    dashboard: str  # "Plan Estratégico" | "Plan de Efectividad" | "Plan de Evaluación de Resultados"
    codigo: str
    nombre: str
    dimension: str = ""          # agrupador principal (Dimensión / Estrategia)
    objetivo: str = ""           # texto del objetivo institucional
    objetivo_num: Optional[int] = None  # 1..7 cuando aplica (Assessment Plan)
    tipo: str = ""               # Directa / Indirecta / Cuantitativa.../ etc
    formula: str = ""
    meta: str = ""
    resultado: str = ""
    evidencia: str = ""
    responsable: str = ""
    estado: str = ESTADO_SIN_DATOS
    estado_original: str = ""
    codigo_relacionado_efectividad: str = ""
    codigo_relacionado_estrategico: str = ""
    oportunidad_mejora: bool = False
    oportunidad_mensaje: str = ""

    def to_dict(self):
        d = self.__dict__.copy()
        return d


# ---------------------------------------------------------------------------
# Objetivos institucionales (IO-1 .. IO-7) — usados por el Plan de Evaluación
# de Resultados a través de la columna "Alineación de IO". Se leen del propio
# Excel (filas de leyenda al final de la hoja) para no hardcodear texto.
# ---------------------------------------------------------------------------

def _read_objetivos_institucionales(ws) -> dict:
    """Lee la tabla de leyenda 'Objetivos institucionales de BGU' al pie de la
    hoja de Evaluación de Resultados. Devuelve {"IO-1": {"nombre":..., "num":1, "desc":...}, ...}
    """
    objetivos = {}
    found_header = False
    for row in ws.iter_rows(values_only=True):
        row_text = " ".join(_clean(c) for c in row if c is not None)
        if "Objetivos institucionales" in row_text:
            found_header = True
            continue
        if not found_header:
            continue
        # columnas relevantes: D = nombre dimensión, G = código IO-x, H = descripción
        nombre = _clean(row[3]) if len(row) > 3 else ""
        codigo_io = _clean(row[6]) if len(row) > 6 else ""
        desc = _clean(row[7]) if len(row) > 7 else ""
        if codigo_io.upper().startswith("IO-"):
            try:
                num = int(re.sub(r"\D", "", codigo_io))
            except ValueError:
                num = None
            objetivos[codigo_io.upper()] = {"nombre": nombre, "num": num, "descripcion": desc}
        elif not codigo_io and not nombre:
            # fila vacía -> fin de la tabla
            if objetivos:
                break
    return objetivos


# ---------------------------------------------------------------------------
# Oportunidades de mejora (Sección 4 del prompt) — lista fija de códigos que
# deben resaltarse con un mensaje especial, según instrucción explícita del
# usuario.
# ---------------------------------------------------------------------------
OPORTUNIDADES_MEJORA = {
    # Plan Estratégico
    "E3-K3": "Puede cambiar a Cumplido si se incorpora la evidencia correspondiente.",
    "E3-K4": "Puede cambiar a Cumplido si se incorpora la evidencia correspondiente.",
    # Plan de Efectividad
    "E3-O01": "Oportunidad de mejora.",
    "E3-O02": "Oportunidad de mejora.",
    # Plan de Evaluación de Resultados
    "D-01": "Podría cambiar de estado cuando exista evidencia suficiente.",
    "D-02": "Podría cambiar de estado cuando exista evidencia suficiente.",
    "D-04": "Podría cambiar de estado cuando exista evidencia suficiente.",
}


# ---------------------------------------------------------------------------
# Parsers específicos por dashboard
# ---------------------------------------------------------------------------

def parse_plan_estrategico(path: str = PATH_ESTRATEGICO) -> list[KPI]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_ESTRATEGICO]

    kpis: list[KPI] = []
    current_dim = ""
    current_obj = ""

    # Datos reales: filas 3 en adelante, columnas A..O (0..14)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        codigo = _clean(row[2]) if len(row) > 2 else ""
        if not codigo or not re.match(r"^E\d+-K\d+$", codigo):
            continue  # fila de totales / vacía / ruido

        dim = _clean(row[0])
        obj = _clean(row[1])
        if dim:
            current_dim = _title_case_es(dim)
        if obj:
            current_obj = obj

        nombre = _clean(row[3])
        cod_efectividad = _clean(row[4])
        meta = _fmt_meta_or_result(row[7])  # Meta 2025-26
        resultado = _fmt_meta_or_result(row[12])  # Resultado 2025-2026
        evidencia = _clean(row[13])
        responsable = _clean(row[11])
        estado_raw = _clean(row[14])
        estado = normalize_estado(estado_raw)

        kpi = KPI(
            dashboard="Plan Estratégico",
            codigo=codigo,
            nombre=nombre,
            dimension=current_dim,
            objetivo=current_obj,
            tipo="",
            formula="",
            meta=meta,
            resultado=resultado,
            evidencia=evidencia,
            responsable=responsable,
            estado=estado,
            estado_original=estado_raw,
            codigo_relacionado_efectividad=cod_efectividad,
        )
        if codigo in OPORTUNIDADES_MEJORA:
            kpi.oportunidad_mejora = True
            kpi.oportunidad_mensaje = OPORTUNIDADES_MEJORA[codigo]
        kpis.append(kpi)

    return kpis


def parse_plan_efectividad(path: str = PATH_EFECTIVIDAD) -> list[KPI]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_EFECTIVIDAD]

    kpis: list[KPI] = []
    current_dim = ""
    current_obj = ""

    # Datos reales: filas 4 en adelante, columnas A..Q (0..16)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        codigo = _clean(row[3]) if len(row) > 3 else ""
        if not codigo or not re.match(r"^E\d+-[IOS]\d*", codigo):
            continue

        dim = _clean(row[0])
        # La columna "Estrategia" viene como "E3. Capital humano administrativo";
        # quitamos el prefijo "E#. " para que la dimensión se lea igual que en
        # el Plan Estratégico.
        dim = re.sub(r"^E\d+\.\s*", "", dim).strip()
        obj = _clean(row[1])
        if dim:
            current_dim = _title_case_es(dim)
        if obj:
            current_obj = obj

        nivel = _clean(row[4])
        nombre = _clean(row[5])
        formula = _clean(row[6])
        meta = _clean(row[7])
        tipo_medida = _clean(row[8])
        resultado = _fmt_meta_or_result(row[13])
        evidencia = _clean(row[14])
        responsable = _clean(row[10])
        estado_raw = _clean(row[16]) if len(row) > 16 else ""
        estado = normalize_estado(estado_raw)

        kpi = KPI(
            dashboard="Plan de Efectividad",
            codigo=codigo,
            nombre=nombre,
            dimension=current_dim,
            objetivo=current_obj,
            tipo=f"{nivel} / {tipo_medida}".strip(" /"),
            formula=formula,
            meta=meta,
            resultado=resultado,
            evidencia=evidencia,
            responsable=responsable,
            estado=estado,
            estado_original=estado_raw,
        )
        if codigo in OPORTUNIDADES_MEJORA:
            kpi.oportunidad_mejora = True
            kpi.oportunidad_mensaje = OPORTUNIDADES_MEJORA[codigo]
        kpis.append(kpi)

    return kpis


def parse_plan_evaluacion(path: str = PATH_EVALUACION) -> tuple[list[KPI], dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_EVALUACION]

    objetivos_map = _read_objetivos_institucionales(ws)

    kpis: list[KPI] = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        codigo = _clean(row[0]) if len(row) > 0 else ""
        if not codigo or not re.match(r"^(D|I)-\d+$", codigo):
            continue

        nombre = _clean(row[1])
        tipo = _clean(row[2])  # Directa / Indirecta
        alineacion_io = _clean(row[3])
        meta = _clean(row[15]) if len(row) > 15 else ""
        responsable = _clean(row[16]) if len(row) > 16 else ""
        resultado = _fmt_meta_or_result(row[17]) if len(row) > 17 else ""
        evidencia = _clean(row[18]) if len(row) > 18 else ""
        estado_raw = _clean(row[19]) if len(row) > 19 else ""
        estado = normalize_estado(estado_raw)
        cod_efectividad = _clean(row[10]) if len(row) > 10 else ""
        cod_estrategico = _clean(row[11]) if len(row) > 11 else ""

        # Alineación de IO puede tener varios códigos IO-x separados por coma
        # o un rango "IO-1–IO-7". Tomamos el PRIMER IO como objetivo primario
        # para efectos de agrupación (se documenta como supuesto en el README).
        first_io_match = re.search(r"IO-?\s?(\d+)", alineacion_io)
        objetivo_num = None
        objetivo_nombre = ""
        if first_io_match:
            io_code = f"IO-{first_io_match.group(1)}"
            objetivo_num = int(first_io_match.group(1))
            objetivo_nombre = objetivos_map.get(io_code, {}).get("nombre", "")

        kpi = KPI(
            dashboard="Plan de Evaluación de Resultados",
            codigo=codigo,
            nombre=nombre,
            dimension=objetivo_nombre,
            objetivo=f"Objetivo {objetivo_num}" if objetivo_num else "",
            objetivo_num=objetivo_num,
            tipo=tipo,
            formula=alineacion_io,
            meta=meta,
            resultado=resultado,
            evidencia=evidencia,
            responsable=responsable,
            estado=estado,
            estado_original=estado_raw,
            codigo_relacionado_efectividad=cod_efectividad,
            codigo_relacionado_estrategico=cod_estrategico,
        )
        if codigo in OPORTUNIDADES_MEJORA:
            kpi.oportunidad_mejora = True
            kpi.oportunidad_mensaje = OPORTUNIDADES_MEJORA[codigo]
        kpis.append(kpi)

    return kpis, objetivos_map


# ---------------------------------------------------------------------------
# Carga global con caché simple en memoria
# ---------------------------------------------------------------------------

@dataclass
class DataBundle:
    estrategico: list[KPI] = field(default_factory=list)
    efectividad: list[KPI] = field(default_factory=list)
    evaluacion: list[KPI] = field(default_factory=list)
    objetivos_map: dict = field(default_factory=dict)
    loaded_at: Optional[datetime] = None

    @property
    def all_kpis(self) -> list[KPI]:
        return self.estrategico + self.efectividad + self.evaluacion


_cache: Optional[DataBundle] = None


def load_all_data(force_reload: bool = False) -> DataBundle:
    global _cache
    if _cache is not None and not force_reload:
        return _cache

    estrategico = parse_plan_estrategico()
    efectividad = parse_plan_efectividad()
    evaluacion, objetivos_map = parse_plan_evaluacion()

    _cache = DataBundle(
        estrategico=estrategico,
        efectividad=efectividad,
        evaluacion=evaluacion,
        objetivos_map=objetivos_map,
        loaded_at=datetime.now(),
    )
    return _cache


DASHBOARD_NAMES = {
    "estrategico": "Plan Estratégico",
    "efectividad": "Plan de Efectividad",
    "evaluacion": "Plan de Evaluación de Resultados",
}
